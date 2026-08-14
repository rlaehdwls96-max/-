"""
REPEAT(BR/PT/SL).xlsx 원본을 직접 읽어서, 대시보드에서 바로 쓸 수 있는
"품번(+색상)별 1행" 와이드 포맷과 "세트비" 테이블을 만든다.

기존 pipelines/stage_repeat_*.py(장부용 long format, fact_dept_metric으로 감)와는
별개로, 대시보드 전용 뷰를 만드는 용도. 원본 셀 구조를 그대로 반영:

- 시트명 = "{카테고리}({YYMM})" 형태 (예: "중요품번(2603)", "중점 SET(2604)")
  단, "BR SET비 계산", "Sheet1"은 제외
- 시트 A1 부근 텍스트로 가격군(기간물/저가) 판별
- 데이터는 (품번, 색상) 쌍마다 2행 1세트: "전년실매출" 행 + "추가 생산" 행
  (현재고는 "전년실매출" 행에만 값이 있고, 품번/색상은 블록 시작 행에만 있어
   이후 행은 이어받는다 — forward-fill)
- 헤더 행("품 번"/"구분" 같은 리터럴 문자열)이 시트 중간에도 반복해서 나오므로
  건너뛴다.

발주 리드타임(2026-08 확정):
- 브라(BR)는 생산 공정상 최소 3개월, 팬티(PT)는 최소 2개월을 확보해야 발주~입고 공백 없이
  회전한다는 실제 업무 기준을 반영해, 브랜드별로 다른 리드타임을 SKU마다 붙여준다.
  SL(임시 2개월)은 실제 기준 확인되는 대로 LEAD_TIME_MONTHS만 고치면 된다.
"""

import re
import sqlite3
from pathlib import Path

import openpyxl
import pandas as pd

MONTH_HEADER_COL = 8  # 0-based: '구분' 다음부터 월별 컬럼 시작 (col index 9)
COL_DESC = 4
COL_SKU = 5
COL_COLOR = 6
COL_STOCK = 7
COL_KIND = 8
COL_MONTH_START = 9

BRAND_MAP = {"REPEAT_BR": "BR", "REPEAT_PT": "PT", "REPEAT_SL": "SL"}

# --- 발주 리드타임 설정 (브랜드별 최소 확보 개월) --------------------------
# 브라는 생산 리드타임이 길어 최소 3개월, 팬티는 최소 2개월 필요.
# SL은 정확한 기준 확인 전까지 PT와 동일한 2개월을 임시로 적용.
LEAD_TIME_MONTHS = {"BR": 3.0, "PT": 2.0, "SL": 2.0}
DEFAULT_LEAD_TIME_MONTHS = 2.0  # 매핑에 없는 브랜드가 나올 경우의 안전값
SAFETY_MARGIN_MONTHS = 0.5  # 리드타임 위에 추가로 확보할 여유분(개월)


def _parse_price_tier(ws) -> str:
    """시트 상단 '* 기간물 (REPEAT 계획)' / '* 저가 (REPEAT 계획)' 텍스트에서 가격군 추출."""
    for row in ws.iter_rows(min_row=1, max_row=2, values_only=True):
        for cell in row:
            if isinstance(cell, str) and cell.strip().startswith("*"):
                if "저가" in cell:
                    return "저가"
                if "기간물" in cell:
                    return "기간물"
                return cell.strip()
    return "미상"


def _parse_sheet_name(sheet_name: str) -> tuple[str, str]:
    """'중요품번(2603)' -> ('중요품번', '2603'). 괄호 없으면 (전체이름, '')."""
    m = re.match(r"^(.*?)\s*\((\d{4})\)", sheet_name)
    if m:
        return m.group(1).strip(), m.group(2)
    return sheet_name.strip(), ""


def parse_repeat_workbook(path: str, brand: str) -> pd.DataFrame:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    records = []

    for sheet_name in wb.sheetnames:
        if sheet_name in ("BR SET비 계산", "Sheet1"):
            continue
        category, season = _parse_sheet_name(sheet_name)
        if not season:
            continue  # 시즌 정보 없는 시트는 건너뜀 (예외적인 시트)

        ws = wb[sheet_name]
        price_tier = _parse_price_tier(ws)

        month_cols: list[str] | None = None
        cur_sku = None
        cur_color = None
        cur_stock = None
        pending_row = None  # {"전년실매출": {...}} 대기 중인 블록

        for row in ws.iter_rows(values_only=True):
            if len(row) <= COL_MONTH_START:
                continue
            sku_cell = row[COL_SKU]
            color_cell = row[COL_COLOR]
            stock_cell = row[COL_STOCK]
            kind_cell = row[COL_KIND]

            # 헤더 반복행: 건너뛰되, 월 컬럼명은 처음 한 번만 기록
            if sku_cell == "품 번" or kind_cell == "구분":
                if month_cols is None:
                    month_cols = [str(c).strip() if c else f"col{i}" for i, c in enumerate(row[COL_MONTH_START:])]
                cur_sku, cur_color, cur_stock = None, None, None
                continue

            if kind_cell not in ("전년실매출", "추가 생산"):
                continue  # 설명 텍스트만 있는 스페이서 행 등

            is_blank_slot = sku_cell is None and color_cell is None and stock_cell is None

            if kind_cell == "전년실매출":
                if is_blank_slot:
                    # 품번당 최대 4색상까지 쓸 수 있는 템플릿인데 실제로 안 쓴 빈 슬롯 —
                    # 이전 색상 값을 이어받으면 안 되는 진짜 빈 행이므로 건너뛴다.
                    pending_row = None
                    continue
                if sku_cell:
                    cur_sku = str(sku_cell).replace("\n", " ").strip()
                if color_cell:
                    cur_color = str(color_cell).strip()
                if stock_cell is not None:
                    cur_stock = stock_cell

                if not cur_sku or not cur_color:
                    continue

                monthly = {}
                if month_cols:
                    for i, col_name in enumerate(month_cols):
                        val = row[COL_MONTH_START + i] if COL_MONTH_START + i < len(row) else None
                        if isinstance(val, (int, float)):
                            monthly[col_name] = val

                pending_row = {
                    "brand": brand, "season": season, "category": category,
                    "price_tier": price_tier, "sku_raw": cur_sku, "color": cur_color,
                    "current_stock": cur_stock, "prev_year_sales_monthly": monthly,
                    "sheet": sheet_name,
                }
            elif kind_cell == "추가 생산":
                if pending_row is None:
                    continue  # 대응하는 전년실매출 행이 없거나(빈 슬롯이라 건너뛴 경우) 무시
                monthly = {}
                if month_cols:
                    for i, col_name in enumerate(month_cols):
                        val = row[COL_MONTH_START + i] if COL_MONTH_START + i < len(row) else None
                        if isinstance(val, (int, float)):
                            monthly[col_name] = val
                pending_row["additional_production_monthly"] = monthly
                records.append(pending_row)
                pending_row = None

    wb.close()

    if not records:
        return pd.DataFrame()

    df = pd.DataFrame(records)

    # 품번 텍스트에서 실제 코드와 컵 범위 분리 (예: "VBRS119 A-G" -> "VBRS119", "A-G")
    def split_sku(s: str) -> tuple[str, str]:
        parts = s.split()
        if len(parts) >= 2 and re.match(r"^[A-Z]-[A-Z]$", parts[-1]):
            return " ".join(parts[:-1]), parts[-1]
        return s, ""

    split = df["sku_raw"].apply(split_sku)
    df["품번"] = split.apply(lambda t: t[0])
    df["컵범위"] = split.apply(lambda t: t[1])

    # 월별 딕셔너리 -> 합계/최근3개월평균 계산
    def monthly_sum(d: dict) -> float:
        return sum(v for v in d.values() if isinstance(v, (int, float)))

    def monthly_recent_avg(d: dict, n: int = 3) -> float:
        vals = [v for v in d.values() if isinstance(v, (int, float))]
        if not vals:
            return 0.0
        recent = vals[-n:] if len(vals) >= n else vals
        return sum(recent) / len(recent)

    df["전년실매출_연합계"] = df["prev_year_sales_monthly"].apply(monthly_sum)
    df["전년실매출_최근3개월평균"] = df["prev_year_sales_monthly"].apply(monthly_recent_avg)
    df["추가생산_합계"] = df["additional_production_monthly"].apply(monthly_sum)

    df = df.drop(columns=["prev_year_sales_monthly", "additional_production_monthly", "sku_raw"])

    # --- 브랜드별 발주 리드타임 부여 ---------------------------------------
    # 브라(BR)는 최소 3개월, 팬티(PT)는 최소 2개월 확보해야 발주~입고 공백이 없다는
    # 실제 업무 기준. 대시보드는 이 컬럼을 그대로 읽어 소진예상개월과 비교한다.
    df["리드타임_개월"] = df["brand"].map(LEAD_TIME_MONTHS).fillna(DEFAULT_LEAD_TIME_MONTHS)
    df["발주기준_개월"] = df["리드타임_개월"] + SAFETY_MARGIN_MONTHS

    # 원본 엑셀 자체에 동일 (브랜드+시즌+카테고리+품번+색상)이 두 번 입력된 경우가 있음
    # (예: VBRQ451/BK가 같은 시트에 실수로 중복 기재, 두번째는 현재고 비어있음).
    # 현재고 값이 있는 행을 우선하고, 나머지는 버린다.
    df["_has_stock"] = df["current_stock"].notna().astype(int)
    df = (
        df.sort_values("_has_stock", ascending=False)
        .drop_duplicates(subset=["brand", "season", "category", "품번", "color"], keep="first")
        .drop(columns=["_has_stock"])
        .reset_index(drop=True)
    )

    return df


def parse_set_ratio_sheet(path: str) -> pd.DataFrame:
    """'BR SET비 계산' 시트: 세트(브라+팬티) 그룹별 판매비/재고비."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "BR SET비 계산" not in wb.sheetnames:
        wb.close()
        return pd.DataFrame()

    ws = wb["BR SET비 계산"]
    records = []
    group_id = 0
    in_group = False

    for row in ws.iter_rows(values_only=True):
        if len(row) <= 15:
            continue
        sku = row[5]
        color = row[6]
        if sku == "품번":
            group_id += 1
            in_group = True
            continue
        if sku in (None, "합계"):
            in_group = False
            continue
        if not in_group or not isinstance(sku, str):
            continue

        records.append({
            "세트그룹": group_id,
            "품번": str(sku).strip(),
            "색상": str(color).strip() if color else "",
            "실판매": row[7],
            "판매비": row[8],
            "현재고": row[9],
            "추가생산": row[10],
            "추가생산_조정": row[11],
            "총재고": row[12],
            "재고비": row[13],
            "1년실판매": row[14],
        })

    wb.close()
    return pd.DataFrame(records)


def build_dashboard_tables(uploads: dict[str, str], out_db: str) -> None:
    """uploads: {'BR': path, 'PT': path, 'SL': path} -> SQLite에 dash_repeat_sku / dash_set_ratio 저장."""
    all_dfs = []
    for brand, path in uploads.items():
        df = parse_repeat_workbook(path, brand)
        if not df.empty:
            all_dfs.append(df)
    repeat_df = pd.concat(all_dfs, ignore_index=True) if all_dfs else pd.DataFrame()

    set_ratio_df = pd.DataFrame()
    if "PT" in uploads:
        set_ratio_df = parse_set_ratio_sheet(uploads["PT"])

    conn = sqlite3.connect(out_db)
    if not repeat_df.empty:
        repeat_df.to_sql("dash_repeat_sku", conn, if_exists="replace", index=False)
    if not set_ratio_df.empty:
        set_ratio_df.to_sql("dash_set_ratio", conn, if_exists="replace", index=False)
    conn.commit()
    conn.close()

    print(f"[parse_repeat] dash_repeat_sku {len(repeat_df)}행, dash_set_ratio {len(set_ratio_df)}행 -> {out_db}")


if __name__ == "__main__":
    import sys
    uploads = {
        "BR": sys.argv[1] if len(sys.argv) > 1 else "REPEAT_BR.xlsx",
        "PT": sys.argv[2] if len(sys.argv) > 2 else "REPEAT_PT.xlsx",
        "SL": sys.argv[3] if len(sys.argv) > 3 else "REPEAT_SL.xlsx",
    }
    out_db = sys.argv[4] if len(sys.argv) > 4 else "data/warehouse.db"
    build_dashboard_tables(uploads, out_db)
